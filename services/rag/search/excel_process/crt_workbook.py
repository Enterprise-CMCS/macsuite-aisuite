import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from common.utils.logger import log

HEADER_ROW = 7
FIRST_DATA_ROW = 8

TIPS_MARKER = "tips and additional"

MAX_SCAN_ROW = 2000

ITEM_COL = 1
LEGAL_CITE_COL = 2
REQUIREMENT_COL = 3
TIP_COL = 4
STATUS_COL = 5
WHERE_FOUND_COL = 6
FOLLOW_UP_COL = 7
COMMENTS_COL = 8


DEFAULT_STATUS_OPTIONS = ["Met", "Unsure", "Not met", "N/A"]

STATUS_WORDING = {
    "MET": ("met", "yes"),
    "NOT MET": ("not met", "no"),
    "UNCLEAR": ("unsure", "unclear"),
}


STANDARD_WORDING = {"MET": "Met", "NOT MET": "Not met", "UNCLEAR": "Unsure"}

MAX_CELL_CHARS = 4000

ANALYSIS_SHEET = "RAG Analysis"
ANALYSIS_COLUMNS = [
    ("Sheet", 26), ("Row", 6), ("Item", 14), ("Legal Cite", 20), ("Requirement", 70),
    ("Status", 10), ("Argument", 70), ("Counter-argument", 70), ("Follow-up Required", 45),
    ("Missing Information", 45), ("Table of Contents Section", 40), ("Where Found", 40),
    ("LLM Confidence", 14), ("Retrieval Confidence", 18), ("Combined Confidence", 18),
    ("Quotes Verified", 14), ("Chunks Retrieved", 14), ("Evidence Quotes", 80),
    ("Model", 26), ("Error", 30),
]

_CELL_RANGE = re.compile(r"^\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?$")


def _text(value):
    if value is None:
        return ""
    return str(value).strip()


def _clip(value):
    text = _text(value)
    return text if len(text) <= MAX_CELL_CHARS else text[:MAX_CELL_CHARS - 3] + "..."


class CRTWorkbook:
    """One Contract Review Tool workbook, open for reading and writing."""

    def __init__(self, path):
        self.path = Path(path)

        self.workbook = load_workbook(self.path, keep_vba=self.path.suffix.lower() == ".xlsm",
                                      data_only=False)
        log.info(f"CRTWorkbook() Opened {self.path.name} with {len(self.workbook.sheetnames)} sheets")

    @property
    def requirement_sheets(self):
        """The sheets that hold requirements, spotted by their header row.

        Matching on the header rather than a hard-coded list of the eleven sheet
        names, because the CRT gets re-issued and the sheets get renamed.
        """
        return [sheet for sheet in self.workbook.worksheets
                if _text(sheet.cell(row=HEADER_ROW, column=REQUIREMENT_COL).value) == "Requirement"]

    def last_reviewable_row(self, sheet):
        limit = min(sheet.max_row, MAX_SCAN_ROW)
        for row in range(FIRST_DATA_ROW, limit + 1):
            if _text(sheet.cell(row=row, column=ITEM_COL).value).lower().startswith(TIPS_MARKER):
                return row - 1
        return limit

    def requirements(self, sheet_names=None, skip_answered=False):
        """Every requirement row in the workbook, in the order a reviewer works them."""
        rows = []
        for sheet in self.requirement_sheets:
            if sheet_names and sheet.title not in sheet_names:
                continue

            for row in range(FIRST_DATA_ROW, self.last_reviewable_row(sheet) + 1):
                item = _text(sheet.cell(row=row, column=ITEM_COL).value)
                requirement = _text(sheet.cell(row=row, column=REQUIREMENT_COL).value)

                # A row with wording but no item number is a section heading.
                if not item or not requirement:
                    continue
                if skip_answered and _text(sheet.cell(row=row, column=STATUS_COL).value):
                    continue

                rows.append({
                    "sheet": sheet.title,
                    "row": row,
                    "item": item,
                    "legal_cite": _text(sheet.cell(row=row, column=LEGAL_CITE_COL).value),
                    "requirement": requirement,
                    "tip": _text(sheet.cell(row=row, column=TIP_COL).value),
                })

        log.info(f"requirements() Collected {len(rows)} requirement rows from {self.path.name}")
        return rows

    def status_options(self, sheet, row):
        """The wording this row's Status dropdown will accept.

        Most rows point at a small range of hidden cells on the same sheet. Rows
        on J. General Terms and Conditions go through
        INDIRECT(SUBSTITUTE($Q..," ","")), which resolves at recalculation time to
        the standard four options, so there is nothing to read there.
        """
        for validation in sheet.data_validations.dataValidation:
            if validation.type != "list":
                continue
            if f"{get_column_letter(STATUS_COL)}{row}" not in validation.sqref:
                continue

            reference = str(validation.formula1 or "").lstrip("=").strip()
            match = _CELL_RANGE.match(reference)
            if not match:
                return list(DEFAULT_STATUS_OPTIONS)

            cells = sheet[reference.replace("$", "")]
            if not isinstance(cells, tuple):
                cells = ((cells,),)
            options = [_text(cell.value) for line in cells for cell in line if _text(cell.value)]
            return options or list(DEFAULT_STATUS_OPTIONS)

        return list(DEFAULT_STATUS_OPTIONS)

    def write_review(self, review):
        """Put one finding into the four columns the reviewer works from."""
        sheet = self.workbook[review.sheet]
        row = review.row
        options = self.status_options(sheet, row)
        wording = match_status(review.status, options)

        follow_up = review.follow_up or review.missing_information
        if wording is None:
            # A. Completeness has no "unsure" option, so an UNCLEAR verdict has
            # nowhere to go. Leaving Status empty is honest - the reviewer still
            # has to answer it - and the reason goes in the comment.
            follow_up = follow_up or "Automated review could not determine this from the contract text."
            log.debug(f"write_review() {review.sheet} row {row}: {review.status} has no matching option "
                      f"in {options}, leaving Status blank")
        else:
            sheet.cell(row=row, column=STATUS_COL).value = wording

        # None rather than "" so a column we have nothing to say about stays blank
        # instead of holding an empty string the reviewer cannot see.
        # Recomputed from the evidence rather than read from review.sources, so
        # re-rendering an older sidecar picks up how we format citations today.
        sheet.cell(row=row, column=WHERE_FOUND_COL).value = _clip(review.where_found()) or None
        sheet.cell(row=row, column=FOLLOW_UP_COL).value = _clip(follow_up) or None
        sheet.cell(row=row, column=COMMENTS_COL).value = \
            _clip(self._comment(review, wording, options)) or None

    def _comment(self, review, wording=None, options=None):
        """Column H: the verdict, the reasoning, the dissent, and how much to trust them."""
        parts = [self._verdict_line(review, wording, options)]
        if review.error:
            parts.append(f"Automated review failed: {review.error}")
        if review.argument:
            parts.append(review.argument)
        if review.counter_argument:
            parts.append(f"Counter-argument: {review.counter_argument}")

        scores = []
        if review.llm_confidence is not None:
            scores.append(f"LLM {review.llm_confidence:.2f}")
        if review.retrieval_confidence is not None:
            scores.append(f"retrieval {review.retrieval_confidence:.2f}")
        if review.combined_confidence is not None:
            scores.append(f"combined {review.combined_confidence:.2f}")
        if scores:
            parts.append("Confidence: " + ", ".join(scores))

        if review.evidence and not review.quotes_verified:
            parts.append("Note: at least one quote could not be matched back to the retrieved text.")

        quotes = format_quotes(review)
        if quotes:

            parts.append(f"Cited contract text:\n{quotes}")

        return "\n\n".join(parts)

    @staticmethod
    def _verdict_line(review, wording, options):

        verdict = STANDARD_WORDING.get(review.status, review.status)
        line = f"Automated status: {verdict}"
        if wording is None:
            # No slot in this sheet's dropdown for the verdict, so column E is blank
            # and this line is the only record of it in the reviewer's columns.
            offered = " / ".join(options or DEFAULT_STATUS_OPTIONS)
            return (f"{line}, on the {' / '.join(DEFAULT_STATUS_OPTIONS)} scale the other sheets use. "
                    f"This sheet's Status dropdown offers {offered}, which has no equivalent, "
                    f"so Status is left for the reviewer to answer.")
        if wording.strip().lower() != verdict.lower():
            return f"{line}, entered as \"{wording}\" in this sheet's wording"
        return line

    def write_analysis(self, reviews):
        """Rebuild the RAG Analysis sheet from scratch so re-runs stay clean."""
        if ANALYSIS_SHEET in self.workbook.sheetnames:
            del self.workbook[ANALYSIS_SHEET]

        sheet = self.workbook.create_sheet(ANALYSIS_SHEET)
        sheet.append([header for header, _ in ANALYSIS_COLUMNS])
        for column, (_, width) in enumerate(ANALYSIS_COLUMNS, start=1):
            sheet.cell(row=1, column=column).font = Font(bold=True)
            sheet.column_dimensions[get_column_letter(column)].width = width

        for review in reviews:
            sheet.append([
                review.sheet, review.row, review.item, review.legal_cite, _clip(review.requirement),
                review.status, _clip(review.argument), _clip(review.counter_argument),
                _clip(review.follow_up), _clip(review.missing_information),
                _clip(review.toc_sections), _clip(review.where_found()),
                review.llm_confidence, review.retrieval_confidence, review.combined_confidence,

                ("Yes" if review.quotes_verified else "No") if review.evidence else None,
                review.chunks_retrieved,
                _clip(format_quotes(review)), review.model, review.error,
            ])

        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        log.info(f"write_analysis() Wrote {len(reviews)} rows to the {ANALYSIS_SHEET} sheet")

    def save(self, output_path):
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            self.workbook.save(output_path)
        except PermissionError:

            fallback = _free_name(output_path)
            log.warning(f"save() {output_path.name} is locked, most likely open in Excel. "
                        f"Saving to {fallback.name} instead.")
            self.workbook.save(fallback)
            output_path = fallback
        log.info(f"save() Saved the reviewed workbook to {output_path}")
        return output_path


def _free_name(path):
    suffix = 2
    while True:
        candidate = path.with_name(f"{path.stem}-{suffix}{path.suffix}")
        if not candidate.exists():
            return candidate
        suffix += 1


def match_status(status, options):
    """The dropdown wording that means `status`, or None if the sheet has no slot for it."""
    wanted = STATUS_WORDING.get(status, ())
    for option in options:
        if option.strip().lower() in wanted:
            return option
    return None


def format_quotes(review):

    lines = []
    for record in review.evidence:
        where = " ".join(part for part in (record.toc_path, record.toc_title) if part).strip()
        location = ", ".join(part for part in (where, record.printed_page) if part)
        location = location or record.citation or record.doc_id
        flag = "" if record.verified else " [unverified]"
        lines.append(f"{location}{flag}\n\"{record.quote}\"".strip())
    return "\n\n".join(lines)
