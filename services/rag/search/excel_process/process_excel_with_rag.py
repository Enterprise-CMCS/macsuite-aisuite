#!/usr/bin/env python3

import argparse
import asyncio
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from common.utils.logger import log
from search.database_searching.review_models import RequirementReview, page_label

DESCRIPTION = "Review every requirement in the CRT workbooks dropped into output_excel/."

HEADER_ROW = 7
FIRST_DATA_ROW = 8

TIPS_MARKER = "tips and additional"

MAX_SCAN_ROW = 2000

ITEM_COL = 1
LEGAL_CITE_COL = 2
REQUIREMENT_COL = 3
TIP_COL = 4
STATUS_COL = 5
WHERE_FOUND_COL = 6
FOLLOW_UP_COL = 7
COMMENTS_COL = 8

DEFAULT_STATUS_OPTIONS = ["Met", "Unsure", "Not met", "N/A"]

STATUS_WORDING = {
    "MET": ("met", "yes"),
    "NOT MET": ("not met", "no"),
    "UNCLEAR": ("unsure", "unclear"),
}

STANDARD_WORDING = {"MET": "Met", "NOT MET": "Not met", "UNCLEAR": "Unsure"}

CONFIDENCE_BANDS = ((0.70, "Strong evidence"), (0.45, "Moderate evidence"))
LOWEST_BAND = "Limited evidence"

MAX_CELL_CHARS = 4000
MAX_ERROR_CHARS = 220

QUOTE_HEADING = "Cited contract text:"

ANALYSIS_SHEET = "RAG Analysis"
ANALYSIS_COLUMNS = [
    ("Sheet", 26), ("Row", 6), ("Item", 14), ("Legal Cite", 20), ("Requirement", 70),
    ("AI Recommended Status", 22), ("AI Response", 70), ("Follow-up Required", 45),
    ("Missing Information", 45), ("Where Found", 40), ("Page number", 14),
    ("LLM Confidence", 14), ("Retrieval Confidence", 18), ("Combined Confidence", 18),
    ("Quotes Verified", 14), ("Chunks Retrieved", 14), ("Evidence Quotes", 80),
    ("Model", 26), ("Error", 30),
]

_CELL_RANGE = re.compile(r"^\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?$")

WORKBOOK_SUFFIXES = (".xlsm", ".xlsx")

DEFAULT_CONCURRENCY = 4


def _text(value):
    if value is None:
        return ""
    return str(value).strip()


def _clip(value):
    text = _text(value)
    return text if len(text) <= MAX_CELL_CHARS else text[:MAX_CELL_CHARS - 3] + "..."


def _squeeze(text):
    return re.sub(r"\s+", " ", _text(text))


def _summarise_error(error):
    text = _squeeze(error)
    if len(text) <= MAX_ERROR_CHARS:
        return text
    return text[:MAX_ERROR_CHARS].rsplit(" ", 1)[0] + " ..."


def confidence_band(score):
    if score is None:
        return None
    for floor, label in CONFIDENCE_BANDS:
        if score >= floor:
            return label
    return LOWEST_BAND


def _free_name(path):
    suffix = 2
    while True:
        candidate = path.with_name(f"{path.stem}-{suffix}{path.suffix}")
        if not candidate.exists():
            return candidate
        suffix += 1


def match_status(status, options):
    wanted = STATUS_WORDING.get(status, ())
    for option in options:
        if option.strip().lower() in wanted:
            return option
    return None


def fit_quotes(records, budget, elsewhere):
    kept = list(records)
    while kept:
        omitted = len(records) - len(kept)
        note = (f"\n\n({omitted} further quote(s) did not fit, see {elsewhere}.)"
                if omitted else "")
        block = _quote_lines(kept)
        if len(block) + len(note) <= budget:
            return block + note
        kept.pop()
    return ""


def _quote_lines(records):
    lines = []
    for record in records:
        # The file name as well as the page, so a reviewer knows which document to
        # open before turning to the page. Verification status is left out of this
        # text - it still drives the Quotes Verified column on the RAG Analysis
        # sheet, but flagging it here on every quote read as noise more than signal.
        location = ", ".join(part for part in (record.doc_id,
                                               page_label(record.page, record.printed_page)) if part)
        lines.append(f"{location}\n\"{_squeeze(record.quote)}\"".strip())
    return "\n\n".join(lines)


class CRTWorkbook:

    def __init__(self, path):
        self.path = Path(path)
        self.workbook = load_workbook(self.path, keep_vba=self.path.suffix.lower() == ".xlsm",
                                      data_only=False)
        log.info(f"CRTWorkbook() Opened {self.path.name} with {len(self.workbook.sheetnames)} sheets")

    @property
    def requirement_sheets(self):
        return [sheet for sheet in self.workbook.worksheets
                if _text(sheet.cell(row=HEADER_ROW, column=REQUIREMENT_COL).value) == "Requirement"]

    def last_reviewable_row(self, sheet):
        limit = min(sheet.max_row, MAX_SCAN_ROW)
        for row in range(FIRST_DATA_ROW, limit + 1):
            if _text(sheet.cell(row=row, column=ITEM_COL).value).lower().startswith(TIPS_MARKER):
                return row - 1
        return limit

    def requirements(self, sheet_names=None, skip_answered=False):
        rows = []
        for sheet in self.requirement_sheets:
            if sheet_names and sheet.title not in sheet_names:
                continue

            for row in range(FIRST_DATA_ROW, self.last_reviewable_row(sheet) + 1):
                item = _text(sheet.cell(row=row, column=ITEM_COL).value)
                requirement = _text(sheet.cell(row=row, column=REQUIREMENT_COL).value)
                if not item or not requirement:
                    continue
                if skip_answered and _text(sheet.cell(row=row, column=STATUS_COL).value):
                    continue

                rows.append({
                    "sheet": sheet.title,
                    "row": row,
                    "item": item,
                    "legal_cite": _text(sheet.cell(row=row, column=LEGAL_CITE_COL).value),
                    "requirement": requirement,
                    "tip": _text(sheet.cell(row=row, column=TIP_COL).value),
                })

        log.info(f"requirements() Collected {len(rows)} requirement rows from {self.path.name}")
        return rows

    def status_options(self, sheet, row):
        for validation in sheet.data_validations.dataValidation:
            if validation.type != "list":
                continue
            if f"{get_column_letter(STATUS_COL)}{row}" not in validation.sqref:
                continue

            reference = str(validation.formula1 or "").lstrip("=").strip()
            if not _CELL_RANGE.match(reference):
                return list(DEFAULT_STATUS_OPTIONS)

            cells = sheet[reference.replace("$", "")]
            if not isinstance(cells, tuple):
                cells = ((cells,),)
            options = [_text(cell.value) for line in cells for cell in line if _text(cell.value)]
            return options or list(DEFAULT_STATUS_OPTIONS)

        return list(DEFAULT_STATUS_OPTIONS)

    def write_review(self, review):
        sheet = self.workbook[review.sheet]
        row = review.row
        options = self.status_options(sheet, row)
        wording = match_status(review.status, options)

        follow_up = review.in_reviewer_terms(review.follow_up or review.missing_information)
        if wording is None:
            follow_up = follow_up or "Automated review could not determine this from the contract text."
            log.debug(f"write_review() {review.sheet} row {row}: {review.status} has no matching option "
                      f"in {options}, leaving Status blank")
        else:
            sheet.cell(row=row, column=STATUS_COL).value = wording

        comment = self._comment(review, wording)
        sheet.cell(row=row, column=WHERE_FOUND_COL).value = _clip(review.where_found()) or None
        sheet.cell(row=row, column=FOLLOW_UP_COL).value = _clip(follow_up) or None
        sheet.cell(row=row, column=COMMENTS_COL).value = _clip(comment) or None

    def _comment(self, review, wording=None):
        parts = [self._verdict_line(review, wording)]
        if review.error:
            parts.append(f"Automated review failed: {_summarise_error(review.error)}")
        if review.argument:
            parts.append(f"AI response: {review.in_reviewer_terms(review.argument)}")

        pages = review.page_numbers()
        if pages:
            parts.append(f"Page number: {pages}")

        score = review.combined_confidence
        if score is None:
            score = review.retrieval_confidence
        band = confidence_band(score)
        if band:
            parts.append(f"Confidence: {band}")

        if not review.evidence and not review.error:
            parts.append("Note: no contract text was quoted for this status. The confidence above "
                         "comes from the passages the search returned, not from a cited provision.")

        comment = "\n\n".join(parts)
        quotes = self._quote_block(review, len(comment) + 2)
        return f"{comment}\n\n{quotes}" if quotes else comment

    @staticmethod
    def _quote_block(review, used):
        budget = MAX_CELL_CHARS - used - len(QUOTE_HEADING) - 1
        block = fit_quotes(review.evidence, budget, f"the {ANALYSIS_SHEET} sheet")
        return f"{QUOTE_HEADING}\n{block}" if block else ""

    @staticmethod
    def _verdict_line(review, wording):
        verdict = STANDARD_WORDING.get(review.status, review.status)
        line = f"AI recommended status: {verdict}"
        if wording is None:
            return (f"{line}, which this sheet's Status dropdown has no option for, "
                    "so Status is left blank for the reviewer to answer.")
        if wording.strip().lower() != verdict.lower():
            return f"{line}, entered as \"{wording}\" in this sheet's wording"
        return line

    def write_analysis(self, reviews):
        if ANALYSIS_SHEET in self.workbook.sheetnames:
            del self.workbook[ANALYSIS_SHEET]

        sheet = self.workbook.create_sheet(ANALYSIS_SHEET)
        sheet.append([header for header, _ in ANALYSIS_COLUMNS])
        for column, (_, width) in enumerate(ANALYSIS_COLUMNS, start=1):
            sheet.cell(row=1, column=column).font = Font(bold=True)
            sheet.column_dimensions[get_column_letter(column)].width = width

        for review in reviews:
            quotes_verified = None
            if review.evidence:
                quotes_verified = "Yes" if review.quotes_verified else "No"

            sheet.append([
                review.sheet, review.row, review.item, review.legal_cite, _clip(review.requirement),
                review.status, _clip(review.in_reviewer_terms(review.argument)),
                _clip(review.in_reviewer_terms(review.follow_up)),
                _clip(review.in_reviewer_terms(review.missing_information)),
                _clip(review.where_found()), review.page_numbers(),
                review.llm_confidence, review.retrieval_confidence, review.combined_confidence,
                quotes_verified,
                review.chunks_retrieved,
                fit_quotes(review.evidence, MAX_CELL_CHARS, "the .reviewed.jsonl sidecar"),
                review.model, review.error,
            ])

        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        log.info(f"write_analysis() Wrote {len(reviews)} rows to the {ANALYSIS_SHEET} sheet")

    def save(self, output_path):
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            self.workbook.save(output_path)
        except PermissionError:
            fallback = _free_name(output_path)
            log.warning(f"save() {output_path.name} is locked, most likely open in Excel. "
                        f"Saving to {fallback.name} instead.")
            self.workbook.save(fallback)
            output_path = fallback
        log.info(f"save() Saved the reviewed workbook to {output_path}")
        return output_path


def default_folder():
    return Path(__file__).resolve().parents[4] / "output_excel"


def find_workbooks(folder):
    return sorted(path for path in folder.glob("*")
                  if path.suffix.lower() in WORKBOOK_SUFFIXES
                  and not path.name.startswith("~$")
                  and ".reviewed" not in path.name)


def saveable(path):
    if not path.exists():
        return True
    try:
        with path.open("r+b"):
            return True
    except OSError:
        return False


def sidecar_path(workbook_path):
    return workbook_path.with_suffix(workbook_path.suffix + ".reviewed.jsonl")


def output_path(workbook_path):
    return workbook_path.with_name(f"{workbook_path.stem}.reviewed{workbook_path.suffix}")


def load_sidecar(path):
    done = {}
    if not path.exists():
        return done

    with path.open(encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                review = RequirementReview.model_validate_json(line)
            except Exception as lclEx:
                log.debug(f"load_sidecar() Ignoring an unreadable sidecar line: {lclEx}")
                continue
            done[(review.sheet, review.row)] = review

    log.info(f"load_sidecar() Reusing {len(done)} requirement(s) already reviewed in {path.name}")
    return done


def append_sidecar(path, review):
    with path.open("a", encoding="utf-8") as handle:
        handle.write(review.model_dump_json() + "\n")


async def review_all(todo, sidecar, concurrency, challenge):
    from search.database_searching.agents import build_deps, review_requirement

    deps = build_deps()
    gate = asyncio.Semaphore(concurrency)
    counter = {"finished": 0}

    async def run_one(entry):
        async with gate:
            review = await review_requirement(
                entry["requirement"],
                deps=deps,
                sheet=entry["sheet"],
                item=entry["item"],
                legal_cite=entry["legal_cite"],
                row=entry["row"],
                challenge=challenge,
            )
            append_sidecar(sidecar, review)
            counter["finished"] += 1
            print(f"  [{counter['finished']}/{len(todo)}] {entry['sheet']} {entry['item']} "
                  f"-> {review.status}"
                  + (f" (confidence {review.combined_confidence:.2f})"
                     if review.combined_confidence is not None else ""))
            return review

    return await asyncio.gather(*(run_one(entry) for entry in todo))


def report_unreviewed(todo):
    print(f"  {len(todo)} not reviewed yet, left blank:")
    for entry in todo[:10]:
        print(f"    {entry['sheet']} row {entry['row']} {entry['item']}")
    if len(todo) > 10:
        print(f"    ... and {len(todo) - 10} more")


async def review_workbook(workbook_path, sheets=None, limit=None, concurrency=DEFAULT_CONCURRENCY,
                          challenge=True, fresh=False, skip_answered=False, render_only=False):
    workbook = CRTWorkbook(workbook_path)
    pending = workbook.requirements(sheet_names=sheets, skip_answered=skip_answered)
    if limit:
        pending = pending[:limit]

    sidecar = sidecar_path(workbook_path)
    if fresh and sidecar.exists():
        sidecar.unlink()
    done = load_sidecar(sidecar)

    todo = [row for row in pending if (row["sheet"], row["row"]) not in done]
    print(f"{workbook_path.name}: {len(pending)} requirement(s), {len(done)} already reviewed, "
          f"{len(todo)} to go")

    if render_only:
        if todo:
            report_unreviewed(todo)
    else:
        for review in await review_all(todo, sidecar, concurrency, challenge):
            done[(review.sheet, review.row)] = review

    reviews = [done[(row["sheet"], row["row"])] for row in pending
               if (row["sheet"], row["row"]) in done]
    for review in reviews:
        workbook.write_review(review)
    workbook.write_analysis(reviews)

    saved = workbook.save(output_path(workbook_path))
    summarise(workbook_path.name, reviews)
    return saved, reviews


def summarise(name, reviews):
    counts = {}
    for review in reviews:
        counts[review.status] = counts.get(review.status, 0) + 1
    errors = sum(1 for review in reviews if review.error)
    unverified = sum(1 for review in reviews if review.evidence and not review.quotes_verified)

    print(f"\n{name}: {len(reviews)} reviewed at {datetime.now(timezone.utc).isoformat(timespec='seconds')}")
    for status in ("MET", "NOT MET", "UNCLEAR"):
        if counts.get(status):
            print(f"  {status:<8} {counts[status]}")
    if unverified:
        print(f"  {unverified} row(s) cite a quote that could not be matched back to the retrieved text")
    if errors:
        print(f"  {errors} row(s) failed and are marked in the General Comments column")


async def run(args):
    folder = Path(args.folder) if args.folder else default_folder()
    if not folder.is_dir():
        print(f"Folder not found: {folder}")
        return 1

    workbooks = find_workbooks(folder)
    if not workbooks:
        print(f"No .xlsm or .xlsx workbooks in {folder}")
        return 1

    locked = [output_path(path) for path in workbooks if not saveable(output_path(path))]
    if locked:
        print("Close these in Excel first, otherwise the run cannot save its findings:")
        for path in locked:
            print(f"  {path}")
        return 1

    verb = "Rendering" if args.render_only else "Reviewing"
    print(f"{verb} {len(workbooks)} workbook(s) in {folder}\n")
    try:
        for workbook_path in workbooks:
            saved, _ = await review_workbook(
                workbook_path,
                sheets=args.sheet or None,
                limit=args.limit,
                concurrency=args.concurrency,
                challenge=not args.no_challenge,
                fresh=args.fresh,
                skip_answered=args.skip_answered,
                render_only=args.render_only,
            )
            print(f"  written to {saved}\n")
    finally:
        if not args.render_only:
            from data_embeddings_storage.database.connection import close_db
            await close_db()

    return 0


def main():
    sys.stdout.reconfigure(errors="replace")

    parser = argparse.ArgumentParser(description=DESCRIPTION)
    parser.add_argument("--folder", help="Folder holding the workbooks (default: output_excel/)")
    parser.add_argument("--sheet", action="append",
                        help="Only this sheet, repeatable. Default is all requirement sheets.")
    parser.add_argument("--limit", type=int, help="Stop after this many requirements, for a smoke test")
    parser.add_argument("--concurrency", type=int, default=DEFAULT_CONCURRENCY,
                        help=f"Requirements in flight at once (default {DEFAULT_CONCURRENCY})")
    parser.add_argument("--no-challenge", action="store_true",
                        help="Skip the challenger and adjudicator, one pass only")
    parser.add_argument("--skip-answered", action="store_true",
                        help="Leave rows that already have a Status alone")
    parser.add_argument("--fresh", action="store_true", help="Discard the sidecar and review everything again")
    parser.add_argument("--render-only", action="store_true",
                        help="Rebuild the workbook from the sidecar without reviewing anything")
    args = parser.parse_args()

    if args.render_only and args.fresh:
        parser.error("--fresh discards the sidecar that --render-only renders from. Pick one.")

    return asyncio.run(run(args))


if __name__ == "__main__":
    sys.exit(main())
