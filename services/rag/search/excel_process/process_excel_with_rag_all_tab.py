import sys
import json
import asyncio
import re
from pathlib import Path
from typing import Optional
from copy import copy
 
from openpyxl import load_workbook
 
src_path = Path(__file__).parent.parent.parent
sys.path.insert(0, str(src_path))
 
from search.database_searching.agents import search_agent, ChatDeps
from common.utils.logger import get_logger
 
logger = get_logger(__name__)
 
MAX_HEADER_SCAN_ROWS = 25
REQUIREMENT_COL = "Requirement"
 
RESULT_COLUMNS = [
    "AI Recommendation",
    "AI Reasoning",
    "Citation Source",
]
 
 
class ExcelRAGProcessor:
    def __init__(self, excel_file_path: str, sheet_name: Optional[str] = None):
        self.excel_file_path = Path(excel_file_path)
        self.sheet_name = sheet_name
        self.workbook = None
        self.worksheet = None
        self.header_row = None
        self.requirement_col = None
        self.result_cols = {}
        self.processed_count = 0
        self.error_count = 0
 
    def list_sheets(self) -> list:
        wb = load_workbook(self.excel_file_path, read_only=True, keep_vba=True)
        try:
            return wb.sheetnames
        finally:
            wb.close()
 
    def load_workbook_preserve_format(self):
        if not self.excel_file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_file_path}")
 
        self.workbook = load_workbook(self.excel_file_path, keep_vba=True)
 
        if self.sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet not found: {self.sheet_name}")
 
        self.worksheet = self.workbook[self.sheet_name]
        self.header_row, self.requirement_col = self._detect_header_row_and_requirement_col()
 
        self._ensure_result_columns()
 
        logger.info(
            f"Loaded workbook with formatting preserved | Sheet={self.sheet_name} "
            f"| Header row={self.header_row} | Requirement col={self.requirement_col}"
        )
 
    def _detect_header_row_and_requirement_col(self):
        for row_num in range(1, MAX_HEADER_SCAN_ROWS + 1):
            for col_num in range(1, self.worksheet.max_column + 1):
                value = self.worksheet.cell(row=row_num, column=col_num).value
                if str(value).strip() == REQUIREMENT_COL:
                    return row_num, col_num
 
        raise ValueError(
            f"Could not find '{REQUIREMENT_COL}' within first "
            f"{MAX_HEADER_SCAN_ROWS} rows of sheet '{self.sheet_name}'"
        )
 
    def _copy_cell_style(self, source_cell, target_cell):
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
 
    def _ensure_result_columns(self):
        existing_headers = {}
 
        for col_num in range(1, self.worksheet.max_column + 1):
            value = self.worksheet.cell(row=self.header_row, column=col_num).value
            if value:
                existing_headers[str(value).strip()] = col_num
 
        next_col = self.worksheet.max_column + 1
 
        for col_name in RESULT_COLUMNS:
            if col_name in existing_headers:
                self.result_cols[col_name] = existing_headers[col_name]
                continue
 
            col_num = next_col
            next_col += 1
 
            header_cell = self.worksheet.cell(row=self.header_row, column=col_num)
            header_cell.value = col_name
 
            previous_header = self.worksheet.cell(row=self.header_row, column=col_num - 1)
            self._copy_cell_style(previous_header, header_cell)
 
            prev_letter = previous_header.column_letter
            new_letter = header_cell.column_letter
            self.worksheet.column_dimensions[new_letter].width = (
                self.worksheet.column_dimensions[prev_letter].width
            )
 
            self.result_cols[col_name] = col_num

    # def _limit_citation_sources(self, source: str, max_sources: int = 5) -> str:
    #     if not source or source == "N/A":
    #         return source
    #     parts = [part.strip() for part in source.split("|") if part.strip()]
    #     return " | ".join(parts[:max_sources])

    def _limit_citation_sources(self, source: str, max_sources: int = 5, max_pages_per_source: int = 5) -> str:
        if not source or source == "N/A":
            return source
    
        sources = [part.strip() for part in source.split("|") if part.strip()]
        limited_sources = []
    
        for item in sources[:max_sources]:
            match = re.search(r"^(.*?),\s*Pages?\s+(.+)$", item, flags=re.I)
    
            if not match:
                limited_sources.append(item)
                continue
    
            doc_name = match.group(1).strip()
            pages_text = match.group(2).strip()
    
            pages = [page.strip() for page in re.split(r",\s*", pages_text)if page.strip()]
            pages = pages[:max_pages_per_source]

            if len(pages) == 1:
                limited_sources.append(f"{doc_name}, Page {pages[0]}")
            else:
                limited_sources.append(f"{doc_name}, Pages {', '.join(pages)}")
    
        return " | ".join(limited_sources)


    async def process_requirement(self, requirement: str, retry_unclear: bool = True) -> dict:
        try:
            logger.info(f"Processing: {requirement[:80]}...")
 
            deps = ChatDeps(acronyms={})
            agent_response = await search_agent.run(user_prompt=requirement, deps=deps)
 
            response_text = (
                agent_response.output
                if hasattr(agent_response, "output")
                else str(agent_response)
            )
 
            parsed = self._parse_agent_response(response_text)
            recommendation = parsed.get("recommendation", "NOT MET")
 
            if recommendation == "NOT MET" and retry_unclear:
                retry_prompt = (
                    "Based on the available documentation, please determine if the following "
                    "requirement is MET or NOT MET. If there is insufficient information, "
                    "explain what specific information is missing.\n\n"
                    f"Requirement: {requirement}\n\n"
                    "Please provide a clear MET or NOT MET determination with specific evidence."
                )
 
                retry_response = await search_agent.run(user_prompt=retry_prompt, deps=deps)
 
                retry_text = (
                    retry_response.output
                    if hasattr(retry_response, "output")
                    else str(retry_response)
                )
 
                parsed = self._parse_agent_response(retry_text)
 
            return {
                "recommendation": parsed.get("recommendation", "NOT MET"),
                "response": parsed.get("response", "No response"),
                "source": parsed.get("source", "N/A"),
            }
 
        except Exception as e:
            logger.error(f"Error processing requirement: {str(e)}")
            self.error_count += 1
            return {
                "recommendation": "ERROR",
                "response": f"Error processing requirement: {str(e)}",
                "source": "N/A",
            }
 
    def _parse_agent_response(self, response_text: str) -> dict:
        try:
            response_text = re.sub(
                r"<thinking>.*?</thinking>",
                "",
                response_text,
                flags=re.DOTALL,
            )
            response_text = response_text.strip()
 
            json_match = re.search(
                r"\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}",
                response_text,
                re.DOTALL,
            )
 
            if json_match:
                try:
                    result = json.loads(json_match.group())
                except json.JSONDecodeError:
                    result = {}
            else:
                result = {}
 
            recommendation = result.get("Recommendation", result.get("Verdict", "NOT MET"))
            response = result.get("Response", result.get("Reasoning", response_text[:500]))
            source = result.get("Source", "N/A")
            # page = result.get("Page", "")
            # source_with_page = f"{source}: {page}" if page and page != "N/A" else source
            # source_with_page = source
            source_with_page = self._limit_citation_sources(source, max_sources=5)
    
 
            return {
                "recommendation": recommendation.upper() if recommendation else "NOT MET",
                "response": response if response else "No response provided",
                "source": source_with_page,
            }
 
        except Exception as e:
            logger.error(f"Error parsing response: {str(e)}")
            return {
                "recommendation": "NOT MET",
                "response": response_text[:500] if response_text else "Error parsing response",
                "source": "N/A",
            }
 
    async def process_all(
        self,
        max_rows: Optional[int] = None,
        save_every: int = 5,
        output_path: Optional[Path] = None,
        skip_existing: bool = True,
    ):
        rows_to_process = []
 
        for row_num in range(self.header_row + 1, self.worksheet.max_row + 1):
            requirement = self.worksheet.cell(
                row=row_num,
                column=self.requirement_col,
            ).value
 
            if not requirement or not str(requirement).strip():
                continue
 
            if skip_existing:
                recommendation_cell = self.worksheet.cell(
                    row=row_num,
                    column=self.result_cols["AI Recommendation"],
                ).value
 
                if recommendation_cell and str(recommendation_cell).strip():
                    continue
 
            rows_to_process.append(row_num)
 
        if max_rows:
            rows_to_process = rows_to_process[:max_rows]
 
        total = len(rows_to_process)
        logger.info(f"Processing {total} requirements from sheet '{self.sheet_name}'")
 
        for i, row_num in enumerate(rows_to_process, start=1):
            requirement = str(
                self.worksheet.cell(
                    row=row_num,
                    column=self.requirement_col,
                ).value
            ).strip()
 
            logger.info(f"[{i}/{total}] Excel row {row_num}: {requirement[:100]}...")
 
            result = await self.process_requirement(requirement)
 
            self._write_result(row_num, "AI Recommendation", result["recommendation"])
            self._write_result(row_num, "AI Reasoning", result["response"])
            self._write_result(row_num, "Citation Source", result["source"])
 
            self.processed_count += 1
 
            if output_path and self.processed_count % save_every == 0:
                self.save(output_path)
                logger.info(f"Progress saved after {self.processed_count} rows")
 
    def _write_result(self, row_num: int, column_name: str, value: str):
        col_num = self.result_cols[column_name]
        cell = self.worksheet.cell(row=row_num, column=col_num)
        cell.value = value
 
        left_cell = self.worksheet.cell(row=row_num, column=col_num - 1)
        self._copy_cell_style(left_cell, cell)
 
        cell.alignment = copy(cell.alignment)
        cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
 
    def save(self, output_path: Path):
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_path)
 
    def close(self):
        if self.workbook:
            self.workbook.close()
 
 
async def main():
    # data_dir = Path(__file__).parent.parent / "data"
 
    # original_file = (
    #     data_dir
    #     / "Contract Review Tool - CSP Base Contract- Amendment 1 (MC-CRS 5691- 6357)-_5691_1.xlsm"
    # )
    # data_dir = Path(r"C:\Moktari\code\rag_results\state_of_oklahoma\semantic")
    # original_file = data_dir / "Contract Review Tool - CSP Base Contract- Amendment 1 (MC-CRS 5691- 6357)-_5691_1.xlsm"
    data_dir = Path("c:\\Moktari\\code\\rag_results\\state_of_NE\\Semantic")
    original_file = data_dir / "MMCC-19-Comprehensive_Streamlined_CRT-3.5.25.xlsm"

    if not original_file.exists():
        print(f"Excel file not found: {original_file}")
        return 1
 
    output_file = original_file.with_name(f"{original_file.stem}_rag_results.xlsm")
 
    workbook_to_open = output_file if output_file.exists() else original_file
 
    sheet_name = sys.argv[1] if len(sys.argv) > 1 else None
 
    probe = ExcelRAGProcessor(str(workbook_to_open))
    available_sheets = probe.list_sheets()
 
    if sheet_name is None:
        print("Available tabs:")
        for name in available_sheets:
            print(f"  - {name}")
        print()
        print('Usage: python process_excel_with_rag.py "<Tab Name>"')
        return 1
 
    if sheet_name not in available_sheets:
        print(f"Tab '{sheet_name}' not found.")
        print(f"Available tabs: {available_sheets}")
        return 1
 
    processor = ExcelRAGProcessor(str(workbook_to_open), sheet_name=sheet_name)
 
    try:
        processor.load_workbook_preserve_format()
 
        await processor.process_all(
            max_rows=None,
            output_path=output_file,
            skip_existing=False,
        )
 
        processor.save(output_file)
 
        print("PROCESSING COMPLETE")
        print(f"Processed tab: {sheet_name}")
        print(f"Processed rows: {processor.processed_count}")
        print(f"Errors: {processor.error_count}")
        print(f"Output workbook: {output_file}")
 
    finally:
        processor.close()
 
    return 0
 
 
if __name__ == "__main__":
    exit_code = asyncio.run(main())
    sys.exit(exit_code)
